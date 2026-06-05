"""
SmartGPA – Upload Hub Router
Endpoints: /upload/file
Allows lecturers to upload raw grades files, validates headers and rows,
and saves them to mock cloud storage with integration into the Delta pipeline.
"""
import csv
import io
import os
import sys
import uuid
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, status

from app.core.dependencies import require_role
from app.models.schemas import UserOut, UserRole
from app.db.databricks_db import save_uploaded_scores_mock
from app.services.databricks_jobs import (
    DatabricksPipelineError, 
    upload_and_run_pipeline, 
    upload_and_trigger_pipeline,
    get_run_output,
    download_workspace_file
)

router = APIRouter(prefix="/upload", tags=["Upload Hub"])
logger = logging.getLogger("smartgpa.upload")

# Đảm bảo thư mục lưu trữ mock cloud storage tồn tại
MOCK_STORAGE_DIR = os.path.join("storage_mock", "raw", "diem")
os.makedirs(MOCK_STORAGE_DIR, exist_ok=True)


def safe_print(message: str = "") -> None:
    encoding = sys.stdout.encoding or "utf-8"
    sys.stdout.write(str(message).encode(encoding, errors="replace").decode(encoding) + "\n")


def parse_score_list(val: str) -> List[float]:
    """Helper: chuyển đổi chuỗi dấu phẩy thành list float"""
    if not val or val.strip() == "":
        return []
    try:
        # Hỗ trợ phân tách bằng cả dấu phẩy hoặc dấu chấm phẩy
        sep = ";" if ";" in val else ","
        return [float(x.strip()) for x in val.split(sep) if x.strip() != ""]
    except ValueError:
        raise ValueError(f"Không thể chuyển đổi chuỗi điểm '{val}' thành danh sách số.")


def _pad_scores(values: List[float], size: int) -> List[Optional[float]]:
    if not values:
        return [None] * size
    padded: List[Optional[float]] = list(values[:size])
    while len(padded) < size:
        padded.append(values[-1])
    return padded


def build_databricks_pipeline_csv(rows: List[Dict[str, Any]]) -> bytes:
    output = io.StringIO()
    fieldnames = [
        "student_id",
        "student_name",
        "ma_mon",
        "ten_mon",
        "ma_lop_hoc_phan",
        "loai_hoc_phan",
        "so_chi_lt",
        "so_chi_th",
        "thuong_xuyen_1",
        "thuong_xuyen_2",
        "giua_ky",
        "thuc_hanh_1",
        "thuc_hanh_2",
        "thuc_hanh_3",
        "diem_cuoi_ky",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()

    for row in rows:
        loai_hp = row.get("loai_hoc_phan", "ly_thuyet")
        tk_scores = row.get("diem_thong_thuong_list") or []
        th_scores = row.get("diem_thuc_hanh_hien_tai") or []

        if loai_hp == "ly_thuyet":
            so_chi_lt, so_chi_th = 2, 0
            tx1, tx2 = _pad_scores(tk_scores, 2)
            giua_ky = row.get("diem_giua_ky")
            th1, th2, th3 = None, None, None
        elif loai_hp == "thuc_hanh":
            so_chi_lt, so_chi_th = 0, 3
            tx1, tx2, giua_ky = None, None, None
            th1, th2, th3 = _pad_scores(th_scores, 3)
        else:
            so_chi_lt, so_chi_th = 2, 1
            tx1, tx2 = _pad_scores(row.get("diem_thuong_ky_lt_list") or [], 2)
            giua_ky = row.get("diem_giua_ky_lt")
            th_integrated = row.get("diem_thuc_hanh_tich_hop")
            th1, th2, th3 = (
                [th_integrated, th_integrated, th_integrated]
                if th_integrated is not None
                else _pad_scores(th_scores, 3)
            )

        writer.writerow({
            "student_id": row.get("student_id"),
            "student_name": row.get("student_name") or row.get("full_name") or row.get("ho_ten") or "",
            "ma_mon": row.get("ma_mon"),
            "ten_mon": row.get("ten_mon") or "",
            "ma_lop_hoc_phan": row.get("ma_lop_hoc_phan"),
            "loai_hoc_phan": loai_hp,
            "so_chi_lt": so_chi_lt,
            "so_chi_th": so_chi_th,
            "thuong_xuyen_1": tx1,
            "thuong_xuyen_2": tx2,
            "giua_ky": giua_ky,
            "thuc_hanh_1": th1,
            "thuc_hanh_2": th2,
            "thuc_hanh_3": th3,
            "diem_cuoi_ky": row.get("diem_cuoi_ky"),
        })

    return output.getvalue().encode("utf-8")


@router.post(
    "/file",
    status_code=status.HTTP_201_CREATED,
    summary="Tải lên file bảng điểm thô – Chỉ Lecturer hoặc Admin",
    description="""
**[Yêu cầu vai trò: Lecturer hoặc Admin]**

Tải lên bảng điểm lớp học phần dưới định dạng file `.csv` hoặc `.xlsx`.
Hệ thống thực hiện kiểm tra cấu trúc nghiêm ngặt (Fail-Fast Validation):
- Kiểm tra các tiêu đề cột bắt buộc.
- Kiểm tra dải điểm thành phần (phải từ 0.0 đến 10.0).
- Trả về chi tiết lỗi hàng nếu dữ liệu không hợp lệ (đạt TC-02).
    """,
)
async def upload_file(
    file: UploadFile = File(...),
    current_user: UserOut = Depends(require_role(UserRole.LECTURER, UserRole.ADMIN)),
) -> dict:
    # 1. Kiểm tra định dạng file
    filename = file.filename or ""
    file_ext = os.path.splitext(filename)[1].lower()
    
    if file_ext not in [".csv", ".xlsx"]:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Định dạng file không hợp lệ. Chỉ chấp nhận file .csv hoặc .xlsx"
        )

    # Đọc nội dung file
    contents = await file.read()
    
    headers = []
    rows_list = []
    decoded = ""
    ma_lop_hoc_phan_parsed = None

    if file_ext == ".csv":
        try:
            decoded = contents.decode("utf-8-sig")  # utf-8-sig để xử lý BOM nếu file từ Excel CSV
        except UnicodeDecodeError:
            try:
                decoded = contents.decode("latin-1")
            except Exception:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Không thể đọc mã hóa của file. Vui lòng đảm bảo file được lưu ở định dạng UTF-8."
                )
        reader = csv.DictReader(io.StringIO(decoded))
        headers = reader.fieldnames or []
        headers = [h.strip() for h in headers]
        rows_list = list(reader)
    else:
        # Xử lý tệp XLSX
        try:
            import openpyxl
            import re
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            ws = wb.active
            
            # Kiểm tra xem có phải định dạng bảng điểm mẫu của IUH không
            is_iuh_template = False
            ma_mon_parsed = None
            ten_mon_parsed = None
            header_row_idx = None
            
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if r_idx > 12:  # Chỉ quét 12 dòng đầu
                    break
                for c_idx, val in enumerate(row, start=1):
                    if val and isinstance(val, str):
                        if "Lớp học phần:" in val:
                            m = re.search(r"Lớp học phần:\s*\[([^\]]+)\]\s*-\s*([^\(]+)(?:\((.+)\))?", val)
                            if m:
                                ma_mon_parsed = m.group(1).strip()
                                from app.db.real_db import COURSE_ID_MAP
                                ma_mon_parsed = COURSE_ID_MAP.get(ma_mon_parsed, ma_mon_parsed)
                                ten_mon_parsed = m.group(2).strip()
                                if m.group(3):
                                    ma_lop_hoc_phan_parsed = m.group(3).strip()
                                is_iuh_template = True
                        if val.strip() == "Mã sinh viên":
                            header_row_idx = r_idx

            if is_iuh_template and header_row_idx:
                # ─── LOGIC PARSE THEO MẪU IUH ─────────────────────────────────
                header_row = [cell.value for cell in ws[header_row_idx]]
                parent_row = [cell.value for cell in ws[header_row_idx - 1]] if header_row_idx > 1 else []
                
                student_id_col = None
                ho_dem_col = None
                ten_col = None
                lop_hoc_col = None
                
                for c_idx, h in enumerate(header_row):
                    if h:
                        h_str = str(h).strip()
                        if h_str == "Mã sinh viên":
                            student_id_col = c_idx
                        elif h_str == "Họ đệm":
                            ho_dem_col = c_idx
                        elif h_str == "Tên":
                            ten_col = c_idx
                        elif h_str == "Lớp học":
                            lop_hoc_col = c_idx
                
                # Xác định loại môn học
                has_th = False
                for p in parent_row:
                    if p and "Thực hành" in str(p):
                        has_th = True
                        break

                # Kiểm tra xem có dữ liệu điểm thực hành thực tế không
                if has_th:
                    has_actual_th_data = False
                    start_row = header_row_idx + 2
                    for r_idx in range(start_row, ws.max_row + 1):
                        row_vals = [cell.value for cell in ws[r_idx]]
                        if len(row_vals) > 12:
                            for col_idx in [10, 11, 12]:
                                val = row_vals[col_idx]
                                if val is not None and str(val).strip() != "":
                                    has_actual_th_data = True
                                    break
                        if has_actual_th_data:
                            break
                    if not has_actual_th_data:
                        logger.info("Môn học có cột Thực hành nhưng không có dữ liệu điểm thực hành. Chuyển sang ly_thuyet.")
                        has_th = False

                loai_hp = "tich_hop" if has_th else "ly_thuyet"
                
                # Cột điểm theo cấu trúc chuẩn của IUH
                giua_ky_col = 7
                tx_cols = [8, 9]
                th_cols = [10, 11, 12]
                cuoi_ky_col = 14
                
                # Thiết lập giả lập headers để vượt qua validation
                headers = ["student_id", "student_name", "ma_mon", "ten_mon", "ma_lop_hoc_phan", "loai_hoc_phan", "diem_giua_ky", "diem_cuoi_ky", "diem_thong_thuong", "diem_thuc_hanh_hien_tai", "diem_thuc_hanh_tich_hop", "diem_giua_ky_lt", "diem_thuong_ky_lt_list"]
                
                rows_list = []
                start_row = header_row_idx + 2
                for r_idx in range(start_row, ws.max_row + 1):
                    row_vals = [cell.value for cell in ws[r_idx]]
                    if len(row_vals) <= (student_id_col or 0):
                        continue
                    sv_id = row_vals[student_id_col]
                    if not sv_id:
                        continue
                    sv_id_str = str(sv_id).strip()
                    if not sv_id_str.isdigit():
                        continue
                        
                    ho_dem = str(row_vals[ho_dem_col] or "").strip() if ho_dem_col is not None else ""
                    ten = str(row_vals[ten_col] or "").strip() if ten_col is not None else ""
                    full_name = f"{ho_dem} {ten}".strip()
                    lop_hoc = str(row_vals[lop_hoc_col] or ma_lop_hoc_phan_parsed or "").strip() if lop_hoc_col is not None else (ma_lop_hoc_phan_parsed or "")
                    
                    def clean_val(val):
                        if val is None or str(val).strip() == "" or str(val).strip().lower() == "nan":
                            return ""
                        return str(val).strip()
                        
                    gk = clean_val(row_vals[giua_ky_col])
                    tx_scores = [clean_val(row_vals[c]) for c in tx_cols]
                    tx_scores = [s for s in tx_scores if s != ""]
                    th_scores = [clean_val(row_vals[c]) for c in th_cols]
                    th_scores = [s for s in th_scores if s != ""]
                    ck = clean_val(row_vals[cuoi_ky_col])
                    
                    row_dict = {
                        "student_id": sv_id_str,
                        "student_name": full_name,
                        "ma_mon": ma_mon_parsed,
                        "ten_mon": ten_mon_parsed,
                        "ma_lop_hoc_phan": lop_hoc,
                        "loai_hoc_phan": loai_hp,
                    }
                    if loai_hp == "ly_thuyet":
                        row_dict["diem_giua_ky"] = gk
                        row_dict["diem_cuoi_ky"] = ck
                        row_dict["diem_thong_thuong"] = ";".join(tx_scores)
                    elif loai_hp == "thuc_hanh":
                        row_dict["diem_thuc_hanh_hien_tai"] = ";".join(th_scores)
                    else: # tich_hop
                        row_dict["diem_thuc_hanh_tich_hop"] = th_scores[0] if th_scores else ""
                        row_dict["diem_giua_ky_lt"] = gk
                        row_dict["diem_thuong_ky_lt_list"] = ";".join(tx_scores)
                        row_dict["diem_cuoi_ky"] = ck
                        
                    rows_list.append(row_dict)
            else:
                # ─── BẢNG ĐIỂM DẠNG API / CHUẨN ĐƠN GIẢN ──────────────────────
                headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        row_dict = {}
                        for i, h in enumerate(headers):
                            if h and i < len(row):
                                val = row[i]
                                row_dict[h] = str(val).strip() if val is not None else ""
                        rows_list.append(row_dict)
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Không thể đọc file XLSX: {str(e)}"
            )

    # Các trường bắt buộc phải có trong file điểm
    required_cols = ["student_id", "ma_mon", "ma_lop_hoc_phan", "loai_hoc_phan"]
    missing_cols = [col for col in required_cols if col not in headers]
    
    if missing_cols:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Cấu trúc file không hợp lệ. Thiếu các cột bắt buộc: {', '.join(missing_cols)}"
        )

    parsed_rows = []
    errors = []

    # Duyệt từng dòng điểm để thực hiện validate nghiệp vụ (Fail-Fast)
    for idx, row in enumerate(rows_list, start=2):  # Dòng 1 là tiêu đề, nên data bắt đầu từ dòng 2
        try:
            student_id = (row.get("student_id") or "").strip()
            ma_mon = (row.get("ma_mon") or "").strip()
            from app.db.real_db import COURSE_ID_MAP
            ma_mon = COURSE_ID_MAP.get(ma_mon, ma_mon)
            ma_lop_hoc = (row.get("ma_lop_hoc_phan") or "").strip()
            loai_hp = (row.get("loai_hoc_phan") or "").strip().lower()

            if not student_id or not ma_mon or not ma_lop_hoc:
                errors.append(f"Dòng {idx}: Mã SV, Mã môn và Mã lớp học phần không được để trống.")
                continue

            if loai_hp not in ["ly_thuyet", "thuc_hanh", "tich_hop"]:
                errors.append(f"Dòng {idx}: Loại học phần '{loai_hp}' không hợp lệ (chỉ chấp nhận ly_thuyet, thuc_hanh, tich_hop).")
                continue

            # Validate các đầu điểm theo từng loại học phần
            parsed_data: Dict[str, Any] = {
                "student_id": student_id,
                "ma_mon": ma_mon,
                "ma_lop_hoc_phan": ma_lop_hoc,
                "loai_hoc_phan": loai_hp,
                "diem_giua_ky": None,
                "diem_cuoi_ky": None,
                "diem_thong_thuong_list": [],
                "diem_thuc_hanh_hien_tai": [],
                "diem_thuc_hanh_tich_hop": None,
                "diem_thuong_ky_lt_list": [],
                "diem_giua_ky_lt": None,
                "student_name": (row.get("student_name") or row.get("full_name") or row.get("ho_ten") or row.get("name") or "").strip(),
                "ten_mon": (row.get("ten_mon") or row.get("subject_name") or "").strip()
            }

            def check_range(name: str, val_str: Optional[str]) -> Optional[float]:
                if not val_str or val_str.strip() == "":
                    return None
                try:
                    f_val = float(val_str.strip())
                    if not (0.0 <= f_val <= 10.0):
                        raise ValueError()
                    return f_val
                except ValueError:
                    raise ValueError(f"Điểm {name} '{val_str}' không hợp lệ. Phải là số từ 0.0 đến 10.0")

            if loai_hp == "ly_thuyet":
                # Nhận giữa kỳ và thường kỳ (dạng danh sách)
                parsed_data["diem_giua_ky"] = check_range("giữa kỳ", row.get("diem_giua_ky"))
                parsed_data["diem_cuoi_ky"] = check_range("cuối kỳ", row.get("diem_cuoi_ky"))
                
                tk_list_str = row.get("diem_thong_thuong") or row.get("diem_thong_thuong_list") or ""
                if tk_list_str:
                    try:
                        scores = parse_score_list(tk_list_str)
                        for s in scores:
                            if not (0.0 <= s <= 10.0):
                                raise ValueError()
                        parsed_data["diem_thong_thuong_list"] = scores
                    except ValueError:
                        errors.append(f"Dòng {idx}: Danh sách điểm thường kỳ '{tk_list_str}' không hợp lệ.")
                        continue

            elif loai_hp == "thuc_hanh":
                # Nhận danh sách điểm thực hành
                th_list_str = row.get("diem_thuc_hanh_hien_tai") or ""
                if th_list_str:
                    try:
                        scores = parse_score_list(th_list_str)
                        for s in scores:
                            if not (0.0 <= s <= 10.0):
                                raise ValueError()
                        parsed_data["diem_thuc_hanh_hien_tai"] = scores
                    except ValueError:
                        errors.append(f"Dòng {idx}: Danh sách điểm thực hành '{th_list_str}' không hợp lệ.")
                        continue

            elif loai_hp == "tich_hop":
                # Nhận điểm TH tích hợp, thường kỳ LT (list), giữa kỳ LT và cuối kỳ LT
                parsed_data["diem_thuc_hanh_tich_hop"] = check_range("thực hành tích hợp", row.get("diem_thuc_hanh_tich_hop"))
                parsed_data["diem_giua_ky_lt"] = check_range("giữa kỳ lý thuyết", row.get("diem_giua_ky_lt"))
                parsed_data["diem_cuoi_ky"] = check_range("cuối kỳ", row.get("diem_cuoi_ky"))
                
                tk_lt_str = row.get("diem_thuong_ky_lt_list") or row.get("diem_thuong_ky_lt") or ""
                if tk_lt_str:
                    try:
                        scores = parse_score_list(tk_lt_str)
                        for s in scores:
                            if not (0.0 <= s <= 10.0):
                                raise ValueError()
                        parsed_data["diem_thuong_ky_lt_list"] = scores
                    except ValueError:
                        errors.append(f"Dòng {idx}: Danh sách điểm thường kỳ lý thuyết '{tk_lt_str}' không hợp lệ.")
                        continue

            parsed_rows.append(parsed_data)

        except ValueError as val_err:
            errors.append(f"Dòng {idx}: {str(val_err)}")
        except Exception as e:
            errors.append(f"Dòng {idx}: Lỗi xử lý định dạng ({str(e)})")

    # 3. Trả về lỗi nếu có bất kỳ dòng nào không vượt qua kiểm tra
    if errors:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={"message": "Dữ liệu file chứa dòng lỗi.", "errors": errors[:20]}  # Giới hạn 20 lỗi đầu tiên tránh tràn response
        )

    # 4. Tránh ghi đè file trên Cloud Storage bằng cách đặt tên file độc nhất
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    random_uuid = uuid.uuid4().hex[:6]
    ma_lop_safe = ma_lop_hoc_phan_parsed or (parsed_rows[0]["ma_lop_hoc_phan"] if parsed_rows else "unknown")
    unique_filename = f"diem_thao_{ma_lop_safe}_{timestamp}_{random_uuid}{file_ext}"
    saved_path = os.path.join(MOCK_STORAGE_DIR, unique_filename)

    # Lưu file thô vào Mock Storage
    try:
        if file_ext == ".xlsx":
            with open(saved_path, "wb") as f:
                f.write(contents)
        else:
            with open(saved_path, "w", encoding="utf-8") as f:
                f.write(decoded)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Lỗi khi lưu trữ file điểm thô: {str(e)}"
        )

    try:
        db_synced = save_uploaded_scores_mock(parsed_rows)
    except Exception as e:
        logger.exception("Failed to sync uploaded scores into local mock store")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Lỗi khi đồng bộ dữ liệu upload vào mock database: {str(e)}",
        )

    # 5b. Tự động đăng ký phân công GV với môn vừa upload (nếu chưa có)
    from app.db.real_db import ASSIGNMENTS_DB, COURSES_DB
    first_parsed = parsed_rows[0] if parsed_rows else {}
    uploaded_ma_mon = first_parsed.get("ma_mon", "")
    uploaded_lop = ma_lop_safe
    if uploaded_ma_mon:
        # Thêm vào COURSES_DB nếu môn chưa có
        if not any(c.get("id") == uploaded_ma_mon for c in COURSES_DB):
            COURSES_DB.append({
                "id": uploaded_ma_mon,
                "name": first_parsed.get("ten_mon") or f"Môn học {uploaded_ma_mon}",
                "type": first_parsed.get("loai_hoc_phan", "ly_thuyet"),
                "credits": (first_parsed.get("so_chi_lt") or 2) + (first_parsed.get("so_chi_th") or 0),
                "chi_lt": first_parsed.get("so_chi_lt") or 2,
                "chi_th": first_parsed.get("so_chi_th") or 0,
                "faculty_id": "CNTT",
                "major_id": "KHDL",
                "is_compulsory": True,
                "hoc_ky": 1,
            })
        # Thêm phân công nếu chưa có
        gv_id = current_user.lecturer_id or current_user.email
        has_assignment = any(
            a.get("lecturer_id") == gv_id and a.get("ma_mon") == uploaded_ma_mon
            for a in ASSIGNMENTS_DB
        )
        if not has_assignment:
            import uuid as _uuid
            ASSIGNMENTS_DB.append({
                "id": f"asgn_{_uuid.uuid4().hex[:8]}",
                "lecturer_id": gv_id,
                "ma_mon": uploaded_ma_mon,
                "ma_lop": uploaded_lop,
                "hoc_ky": "HKII 2025-2026",
            })
            logger.info(f"Auto-registered assignment: GV {gv_id} → môn {uploaded_ma_mon}")

    databricks_filename = unique_filename.replace(file_ext, "_databricks.csv")
    databricks_payload = build_databricks_pipeline_csv(parsed_rows)

    try:
        databricks_result = upload_and_trigger_pipeline(databricks_filename, databricks_payload)
    except DatabricksPipelineError as e:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Databricks pipeline failed to trigger: {str(e)}",
        )

    # 6. Ghi log hoạt động hệ thống (System Activity Audit Log)
    from app.db.fake_db import ACTIVITY_LOGS
    subject_names = {
        "INT1001": "Lập trình Python (Tích hợp)",
        "INT1002": "Cơ sở dữ liệu",
        "GDQP102": "Giáo dục quốc phòng*",
        "mon_1": "Cấu trúc dữ liệu & Giải thuật",
        "mon_2": "Mạng máy tính",
        "mon_3": "Thực hành Hệ điều hành",
        "mon_4": "Thực hành Lập trình hướng đối tượng"
    }
    
    first_row = parsed_rows[0] if parsed_rows else {}
    actual_mon = first_row.get("ma_mon", ma_lop_safe)
    sub_name = subject_names.get(actual_mon, f"Môn học {actual_mon}")

    new_log = {
        "id": f"act_{uuid.uuid4().hex[:6]}",
        "actor_email": current_user.email,
        "actor_name": current_user.full_name,
        "action": "upload",
        "subject_id": actual_mon,
        "subject_name": sub_name,
        "details": f"Nạp bảng điểm lớp {ma_lop_safe} và đang khởi chạy Databricks pipeline. Tệp tin: {unique_filename}.",
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    ACTIVITY_LOGS.append(new_log)

    # ─── Gửi thông báo in-app & giả lập email cho Sinh viên ───
    from app.db.fake_db import USERS_DB
    from app.db.databricks_db import MOCK_GOLD_DB
    
    unique_student_ids = list(set(r["student_id"] for r in parsed_rows))
    for s_id in unique_student_ids:
        student_user = None
        for u in USERS_DB.values():
            if u.get("role") == "student" and u.get("student_id") == s_id:
                student_user = u
                break
                
        if student_user:
            student_email = student_user["email"]
            student_name = student_user["full_name"]
            
            # Kiểm tra xem có cảnh báo hay nguy cơ không
            rec = MOCK_GOLD_DB.get((s_id, actual_mon))
            is_warn = rec.get("status_canh_bao") == "Nguy co" if rec else False
            
            title = "Cảnh báo học vụ khẩn cấp ⚠️" if is_warn else "Cập nhật bảng điểm môn học 📥"
            msg = (
                f"Điểm số môn {sub_name} ({actual_mon}) đã được cập nhật bởi giảng viên {current_user.full_name}. "
                f"Trạng thái: {'⚠️ NGUY CƠ RỚT MÔN' if is_warn else 'An toàn'}."
            )
            
            noti = {
                "id": f"noti_{uuid.uuid4().hex[:6]}",
                "title": title,
                "message": msg,
                "type": "warning" if is_warn else "upload",
                "sender": current_user.full_name,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "is_read": False
            }
            if "notifications" not in student_user:
                student_user["notifications"] = []
            student_user["notifications"].append(noti)
            
            # In giả lập gửi email về console
            safe_print(f"\n=========================================================================")
            safe_print(f"SENDING AUTO-UPLOAD EMAIL NOTIFICATION TO: {student_email}")
            safe_print(f"SUBJECT: [SmartGPA] {title} - Môn {sub_name}")
            safe_print(f"BODY:")
            safe_print(f"Chào {student_name},")
            safe_print(f"Giảng viên {current_user.full_name} vừa cập nhật bảng điểm môn {sub_name} ({actual_mon}) trên SmartGPA.")
            if is_warn:
                safe_print(f"CẢNH BÁO CỰC KỲ QUAN TRỌNG: Điểm số hiện tại của bạn có nguy cơ rớt môn.")
                safe_print(f"Hãy truy cập ngay SmartGPA để chạy Simulation Engine tính toán điểm thi cuối kỳ tối thiểu cần đạt!")
            else:
                safe_print(f"Điểm số hiện tại của bạn đang ở mức an toàn. Hãy đăng nhập để lập mục tiêu điểm số cuối kỳ.")
            safe_print(f"Trân trọng,")
            safe_print(f"Hệ thống SmartGPA")
            safe_print(f"=========================================================================\n")

    return {
        "message": "Upload thành công và đang khởi chạy Databricks pipeline.",
        "filename": unique_filename,
        "databricks_filename": databricks_filename,
        "records_processed": len(parsed_rows),
        "db_synced": db_synced,
        "storage_destination": databricks_result["csv_path"],
        "workspace_path": databricks_result.get("workspace_path"),
        "pipeline_status": "RUNNING",
        "databricks_run_id": databricks_result["run_id"],
        "log_id": new_log["id"]
    }


@router.get(
    "/status/{run_id}",
    summary="Kiểm tra trạng thái chạy Databricks Pipeline",
    description="Truy vấn trạng thái hiện tại (RUNNING, SUCCESS, FAILED) của run_id trên Databricks.",
)
def get_pipeline_status(
    run_id: int,
    current_user: UserOut = Depends(require_role(UserRole.LECTURER, UserRole.ADMIN)),
) -> dict:
    try:
        from app.services.databricks_jobs import check_run_status
        status_res = check_run_status(run_id)
        
        if status_res.get("status") == "SUCCESS":
            # Đồng bộ kết quả đã làm sạch từ file CSV trên Databricks về database local
            try:
                import json
                import csv
                import io
                from app.services.databricks_jobs import get_run_output, download_workspace_file
                from app.db.databricks_db import MOCK_GOLD_DB, sync_gold_to_silver
                
                run_output = get_run_output(run_id)
                result_str = run_output.get("notebook_output", {}).get("result")
                if result_str:
                    output_data = json.loads(result_str)
                    processed_csv_path = output_data.get("processed_csv_path")
                    if processed_csv_path:
                        csv_content = download_workspace_file(processed_csv_path)
                        
                        f = io.StringIO(csv_content.decode('utf-8', errors='replace'))
                        reader = csv.DictReader(f)
                        rows = list(reader)
                        
                        synced_count = 0
                        for row in rows:
                            student_id = row.get("student_id")
                            ma_mon = row.get("ma_mon")
                            if not student_id or not ma_mon:
                                continue
                                
                            def to_float(v):
                                if v is None or v.strip() == "" or v.strip().lower() == "null" or v.strip().lower() == "none":
                                    return None
                                try:
                                    return float(v.strip())
                                except ValueError:
                                    return None
                                
                            def to_float_list(v):
                                if v is None or v.strip() == "" or v.strip().lower() == "null" or v.strip().lower() == "none":
                                    return []
                                return [float(x) for x in v.split(";") if x.strip() != ""]

                            mock_record = {
                                "student_id": student_id,
                                "student_name": row.get("student_name") or row.get("ho_ten") or "",
                                "ma_mon": ma_mon,
                                "ten_mon": row.get("ten_mon") or "",
                                "ma_lop_hoc_phan": row.get("ma_lop_hoc_phan") or "",
                                "diem_tong_ket": to_float(row.get("diem_tong_ket")) or to_float(row.get("diem_tich_luy_hien_tai")),
                                "diem_chu": row.get("diem_chu") or row.get("diem_chu_hien_tai") or "F",
                                "diem_he_4": to_float(row.get("diem_he_4")) or 0.0,
                                "ket_qua": row.get("ket_qua") or ("Dat" if (row.get("diem_chu_hien_tai") or "F") != "F" else "Khong dat"),
                                "diem_thong_thuong": to_float_list(row.get("diem_thong_thuong")),
                                "diem_giua_ky": to_float(row.get("diem_giua_ky")),
                                "diem_cuoi_ky": to_float(row.get("diem_cuoi_ky")),
                                "loai_hoc_phan": row.get("loai_hoc_phan") or "ly_thuyet",
                                "so_chi_lt": int(to_float(row.get("so_chi_lt")) or 2),
                                "so_chi_th": int(to_float(row.get("so_chi_th")) or 0),
                                "tong_so_chi": int(to_float(row.get("tong_so_chi")) or 2),
                                "diem_thuc_hanh_hien_tai": to_float_list(row.get("diem_thuc_hanh_hien_tai")),
                                "diem_thuc_hanh_tich_hop": to_float(row.get("diem_thuc_hanh_tich_hop")),
                                "diem_thuong_ky_lt_list": to_float_list(row.get("diem_thuong_ky_lt_list")),
                                "diem_giua_ky_lt": to_float(row.get("diem_giua_ky_lt")),
                                "status_canh_bao": "Nguy co" if "CANH BAO" in (row.get("status_canh_bao_final") or "") or "Nguy co" in (row.get("status_canh_bao_final") or "") else "An toan"
                            }
                            
                            MOCK_GOLD_DB[(student_id, ma_mon)] = mock_record
                            synced_count += 1
                        
                        sync_gold_to_silver()
                        try:
                            from app.db.persistence import save_db_to_disk
                            save_db_to_disk()
                        except Exception as save_err:
                            logger.error(f"Lỗi khi lưu database: {save_err}")
                        logger.info(f"Đã đồng bộ thành công {synced_count} dòng dữ liệu từ Databricks CSV về local database.")
            except Exception as sync_err:
                logger.error(f"Lỗi khi đồng bộ kết quả Databricks CSV về DB local: {sync_err}", exc_info=True)
                
        return status_res
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Lỗi khi kiểm tra trạng thái run_id {run_id}: {str(e)}",
        )


@router.get(
    "/activities",
    response_model=List[dict],
    summary="Lấy toàn bộ nhật ký hoạt động hệ thống – Chỉ Admin",
    description="Truy xuất lịch sử tải điểm, chỉnh sửa điểm của toàn bộ giảng viên.",
)
def get_activities(
    current_user: UserOut = Depends(require_role(UserRole.ADMIN)),
) -> List[dict]:
    from app.db.fake_db import ACTIVITY_LOGS
    # Trả về danh sách log đảo ngược (mới nhất lên đầu)
    return list(reversed(ACTIVITY_LOGS))


@router.post(
    "/edit",
    status_code=status.HTTP_200_OK,
    summary="Chỉnh sửa điểm học sinh trực tiếp – Chỉ Lecturer hoặc Admin",
    description="Cho phép Giảng viên hoặc Admin điều chỉnh trực tiếp điểm của học viên trong Delta Gold Table.",
)
def edit_grade(
    req_body: dict,  # Dùng dict nhận linh hoạt
    current_user: UserOut = Depends(require_role(UserRole.LECTURER, UserRole.ADMIN)),
) -> dict:
    from app.db.databricks_db import MOCK_GOLD_DB
    from app.db.fake_db import ACTIVITY_LOGS
    
    student_id = req_body.get("student_id")
    ma_mon = req_body.get("ma_mon")
    
    if not student_id or not ma_mon:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Thiếu tham số student_id hoặc ma_mon."
        )
        
    key = (student_id, ma_mon)
    
    subject_names = {
        "INT1001": "Lập trình Python (Tích hợp)",
        "INT1002": "Cơ sở dữ liệu",
        "GDQP102": "Giáo dục quốc phòng*",
        "mon_1": "Cấu trúc dữ liệu & Giải thuật",
        "mon_2": "Mạng máy tính",
        "mon_3": "Thực hành Hệ điều hành",
        "mon_4": "Thực hành Lập trình hướng đối tượng"
    }
    sub_name = subject_names.get(ma_mon, f"Môn học {ma_mon}")
    
    # 1. Tìm hoặc tạo mới bản ghi điểm trong MOCK_GOLD_DB
    old_details = "Chưa có điểm thành phần"
    if key in MOCK_GOLD_DB:
        record = MOCK_GOLD_DB[key]
        old_details_list = []
        if record.get("diem_giua_ky") is not None:
            old_details_list.append(f"GK={record['diem_giua_ky']}")
        if record.get("diem_thong_thuong"):
            old_details_list.append(f"TK={record['diem_thong_thuong']}")
        if record.get("diem_thuc_hanh_hien_tai"):
            old_details_list.append(f"TH={record['diem_thuc_hanh_hien_tai']}")
        if record.get("diem_thuc_hanh_tich_hop") is not None:
            old_details_list.append(f"TH_TichHop={record['diem_thuc_hanh_tich_hop']}")
        if old_details_list:
            old_details = ", ".join(old_details_list)
    else:
        # Tạo mới bản ghi điểm
        MOCK_GOLD_DB[key] = {
            "student_id": student_id,
            "ma_mon": ma_mon,
            "diem_thong_thuong": [],
            "diem_giua_ky": None,
            "diem_cuoi_ky": None,
            "loai_hoc_phan": "ly_thuyet",
            "so_chi_lt": 2,
            "so_chi_th": 0,
            "tong_so_chi": 2,
            "status_canh_bao": "An toan"
        }
    
    record = MOCK_GOLD_DB[key]
    
    # 2. Cập nhật các trường
    updated_list = []
    
    if "diem_giua_ky" in req_body and req_body["diem_giua_ky"] is not None:
        val = req_body["diem_giua_ky"]
        record["diem_giua_ky"] = float(val)
        updated_list.append(f"GK={val}")
        
    if "diem_thong_thuong_list" in req_body and req_body["diem_thong_thuong_list"] is not None:
        val = req_body["diem_thong_thuong_list"]
        record["diem_thong_thuong"] = [float(x) for x in val]
        updated_list.append(f"TK={val}")
        
    if "diem_thuc_hanh_hien_tai" in req_body and req_body["diem_thuc_hanh_hien_tai"] is not None:
        val = req_body["diem_thuc_hanh_hien_tai"]
        record["diem_thuc_hanh_hien_tai"] = [float(x) for x in val]
        record["loai_hoc_phan"] = "thuc_hanh"
        record["so_chi_lt"] = 0
        record["so_chi_th"] = 3
        record["tong_so_chi"] = 3
        updated_list.append(f"TH={val}")
        
    if "diem_thuc_hanh_tich_hop" in req_body and req_body["diem_thuc_hanh_tich_hop"] is not None:
        val = req_body["diem_thuc_hanh_tich_hop"]
        record["diem_thuc_hanh_tich_hop"] = float(val)
        record["loai_hoc_phan"] = "tich_hop"
        record["so_chi_lt"] = 2
        record["so_chi_th"] = 1
        record["tong_so_chi"] = 3
        updated_list.append(f"TH_TichHop={val}")

    new_details = ", ".join(updated_list) if updated_list else "Không thay đổi điểm"
    
    # Đánh giá cảnh báo học vụ sơ bộ lại
    scores_to_check = []
    if record["diem_giua_ky"] is not None:
        scores_to_check.append(record["diem_giua_ky"])
    if record["diem_thong_thuong"]:
        scores_to_check.extend(record["diem_thong_thuong"])
    if record["diem_thuc_hanh_hien_tai"]:
        scores_to_check.extend(record["diem_thuc_hanh_hien_tai"])
    if record["diem_thuc_hanh_tich_hop"] is not None:
        scores_to_check.append(record["diem_thuc_hanh_tich_hop"])
        
    if scores_to_check and (sum(scores_to_check) / len(scores_to_check)) < 4.0:
        record["status_canh_bao"] = "Nguy co"
    else:
        record["status_canh_bao"] = "An toan"
        
    # 3. Ghi log hoạt động
    log_detail = f"Sửa điểm SV {student_id}. Thay đổi: {new_details} (Trước đó: {old_details})."
    
    new_log = {
        "id": f"act_{uuid.uuid4().hex[:6]}",
        "actor_email": current_user.email,
        "actor_name": current_user.full_name,
        "action": "edit",
        "subject_id": ma_mon,
        "subject_name": sub_name,
        "details": log_detail,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    ACTIVITY_LOGS.append(new_log)
    
    # ─── Gửi thông báo in-app & giả lập email cho Sinh viên khi được sửa điểm ───
    from app.db.fake_db import USERS_DB
    
    student_user = None
    for u in USERS_DB.values():
        if u.get("role") == "student" and u.get("student_id") == student_id:
            student_user = u
            break
            
    if student_user:
        student_email = student_user["email"]
        student_name = student_user["full_name"]
        
        is_warn = record.get("status_canh_bao") == "Nguy co"
        title = "Điểm số thay đổi ✏️"
        msg = (
            f"Điểm số môn {sub_name} ({ma_mon}) của bạn đã được điều chỉnh bởi giảng viên {current_user.full_name}. "
            f"Chi tiết thay đổi: {new_details}."
        )
        
        noti = {
            "id": f"noti_{uuid.uuid4().hex[:6]}",
            "title": title,
            "message": msg,
            "type": "edit",
            "sender": current_user.full_name,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "is_read": False
        }
        if "notifications" not in student_user:
            student_user["notifications"] = []
        student_user["notifications"].append(noti)
        
        # In giả lập email ra console
        safe_print(f"\n=========================================================================")
        safe_print(f"SENDING GRADE ADJUSTMENT EMAIL NOTIFICATION TO: {student_email}")
        safe_print(f"SUBJECT: [SmartGPA] Thay đổi điểm môn {sub_name}")
        safe_print(f"BODY:")
        safe_print(f"Chào {student_name},")
        safe_print(f"Giảng viên {current_user.full_name} đã điều chỉnh điểm môn {sub_name} ({ma_mon}) trên SmartGPA.")
        safe_print(f"Chi tiết thay đổi mới: {new_details} (Trước đó: {old_details}).")
        if is_warn:
            safe_print(f"CẢNH BÁO: Điểm số hiện tại của bạn đang nằm trong diện nguy cơ cảnh báo học vụ.")
        safe_print(f"Trân trọng,")
        safe_print(f"Hệ thống SmartGPA")
        safe_print(f"=========================================================================\n")
    
    return {
        "message": f"Cập nhật điểm thành công cho SV {student_id} môn {ma_mon}.",
        "updated_fields": updated_list,
        "log_id": new_log["id"]
    }
