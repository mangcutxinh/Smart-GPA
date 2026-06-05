from __future__ import annotations
"""
SmartGPA – Lecturer Router
Endpoints cho giảng viên:
- Xem danh sách môn được phân công
- Xem danh sách điểm sinh viên theo môn
- Chỉnh sửa / xóa điểm sinh viên
- Upload bảng điểm (CSV / XLSX)
"""
import csv
import io
import logging
from datetime import datetime
from typing import List, Optional, Dict, Any
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile, Query
from pydantic import BaseModel

from app.core.dependencies import require_role
from app.models.schemas import UserOut, UserRole, AdminGradeUpdate
from app.db.real_db import (
    ASSIGNMENTS_DB, COURSES_DB, USERS_DB,
    CURRENT_SEMESTER, SCORE_HISTORY_DB, ACTIVITY_LOGS
)
from app.db.databricks_db import MOCK_GOLD_DB, sync_gold_to_silver

logger = logging.getLogger("smartgpa.lecturer")
router = APIRouter(prefix="/lecturer", tags=["Lecturer"])
_lecturer_dep = Depends(require_role(UserRole.LECTURER))


def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ── Schemas ───────────────────────────────────────────────────

class GradeEditBody(BaseModel):
    diem_thong_thuong: Optional[List[float]] = None
    diem_giua_ky: Optional[float] = None
    diem_cuoi_ky: Optional[float] = None
    diem_thuc_hanh_hien_tai: Optional[List[float]] = None
    diem_thuc_hanh_tich_hop: Optional[float] = None
    reason: Optional[str] = "Giảng viên cập nhật"


# ── Helpers ───────────────────────────────────────────────────

def map_letter_to_gpa(diem_chu: Optional[str]) -> Optional[float]:
    if not diem_chu:
        return None
    mapping = {
        "A+": 4.0,
        "A": 4.0,
        "B+": 3.5,
        "B": 3.0,
        "C+": 2.5,
        "C": 2.0,
        "D+": 1.5,
        "D": 1.0,
        "F": 0.0
    }
    return mapping.get(diem_chu.upper())


def get_grade_letter(tong: float) -> tuple[str, float]:
    if tong >= 9.0:  return ("A+", 4.0)
    if tong >= 8.5:  return ("A",  4.0)
    if tong >= 8.0:  return ("B+", 3.5)
    if tong >= 7.0:  return ("B",  3.0)
    if tong >= 6.0:  return ("C+", 2.5)
    if tong >= 5.5:  return ("C",  2.0)
    if tong >= 5.0:  return ("D+", 1.5)
    if tong >= 4.0:  return ("D",  1.0)
    return ("F", 0.0)


def _get_lecturer_courses(lecturer_user: UserOut) -> List[dict]:
    """Lấy danh sách môn GV đang phụ trách (từ ASSIGNMENTS_DB)"""
    result = []
    seen = set()
    # ASSIGNMENTS_DB is a list
    assignment_iter = ASSIGNMENTS_DB if isinstance(ASSIGNMENTS_DB, list) else ASSIGNMENTS_DB.values()
    for assign in assignment_iter:
        ma_mon = assign.get("ma_mon", "")
        lec_id = assign.get("lecturer_id", "")
        matched = (
            lec_id == lecturer_user.lecturer_id
            or lec_id == lecturer_user.email
            or lec_id == lecturer_user.id
        )
        if matched and ma_mon not in seen:
            seen.add(ma_mon)
            # COURSES_DB is a list, need to search
            course_info = next((c for c in COURSES_DB if c.get("id") == ma_mon), {})
            student_count = sum(
                1 for k in MOCK_GOLD_DB
                if (isinstance(k, tuple) and k[1] == ma_mon)
            )
            result.append({
                "ma_mon": ma_mon,
                "ten_mon": course_info.get("name", ma_mon),
                "loai": course_info.get("type", "ly_thuyet"),
                "tin_chi": course_info.get("credits", 0),
                "ma_lop": assign.get("ma_lop", ""),
                "hoc_ky": assign.get("hoc_ky", CURRENT_SEMESTER.get("display", "")),
                "so_sinh_vien": student_count,
                "assignment_id": assign.get("id", ""),
            })
    return result


def _get_grades_for_course(ma_mon: str) -> List[dict]:
    """Lấy tất cả bản ghi điểm của môn học"""
    records = []
    
    # 1. Thử lấy từ Cloud Databricks nếu được cấu hình
    from app.db.databricks_db import query_course_grades_from_cloud
    cloud_rows = query_course_grades_from_cloud(ma_mon)
    if cloud_rows is not None:
        for row in cloud_rows:
            sid = row.get("student_id")
            sname = row.get("student_name") or f"Sinh viên {sid}"
            loai_hp = row.get("loai_hoc_phan", "ly_thuyet")
            
            tx1 = row.get("thuong_xuyen_1")
            tx2 = row.get("thuong_xuyen_2")
            tx_list = [x for x in (tx1, tx2) if x is not None]
            
            th1 = row.get("thuc_hanh_1")
            th2 = row.get("thuc_hanh_2")
            th3 = row.get("thuc_hanh_3")
            th_list = [x for x in (th1, th2, th3) if x is not None]
            th_avg = sum(th_list) / len(th_list) if th_list else 0.0
            
            # Tính điểm tổng kết, điểm chữ và cảnh báo đồng bộ với update logic
            ck_val = row.get("diem_cuoi_ky")

            if loai_hp == "ly_thuyet":
                tk_avg = sum(tx_list) / len(tx_list) if tx_list else 0.0
                gk_val = row.get("giua_ky") or 0.0
                
                if ck_val is None:
                    diem_tong_ket = None
                    diem_chu = None
                    status_canh_bao = "An toan" if (0.2 * tk_avg + 0.3 * gk_val) >= 2.0 else "Nguy co"
                else:
                    diem_tong_ket = round(0.2 * tk_avg + 0.3 * gk_val + 0.5 * ck_val, 1)
                    diem_tong_ket = min(10.0, diem_tong_ket)
                    if ck_val < 3.0 or diem_tong_ket < 4.0:
                        diem_chu = "F"
                        status_canh_bao = "Nguy co"
                    else:
                        diem_chu, _ = get_grade_letter(diem_tong_ket)
                        status_canh_bao = "An toan"
            
            elif loai_hp == "thuc_hanh":
                diem_tong_ket = round(th_avg, 1) if th_list else 0.0
                if diem_tong_ket < 4.0:
                    diem_chu = "F"
                    status_canh_bao = "Nguy co"
                else:
                    diem_chu, _ = get_grade_letter(diem_tong_ket)
                    status_canh_bao = "An toan"
            
            else: # tich_hop
                tk_avg = sum(tx_list) / len(tx_list) if tx_list else 0.0
                gk_val = row.get("giua_ky") or 0.0
                chi_lt = row.get("so_chi_lt") or 2
                chi_th = row.get("so_chi_th") or 1
                
                if ck_val is None:
                    diem_tong_ket = None
                    diem_chu = None
                    status_canh_bao = "An toan" if th_avg >= 4.0 else "Nguy co"
                else:
                    lt_score = 0.2 * tk_avg + 0.3 * gk_val + 0.5 * ck_val
                    diem_tong_ket = round((lt_score * chi_lt + th_avg * chi_th) / (chi_lt + chi_th), 1)
                    diem_tong_ket = min(10.0, diem_tong_ket)
                    if ck_val < 3.0 or th_avg < 3.0 or diem_tong_ket < 4.0:
                        diem_chu = "F"
                        status_canh_bao = "Nguy co"
                    else:
                        diem_chu, _ = get_grade_letter(diem_tong_ket)
                        status_canh_bao = "An toan"
                
            records.append({
                "_key": f"{sid}_{ma_mon}",
                "student_id": sid,
                "ten_sv": sname,
                "ma_mon": ma_mon,
                "ten_mon": row.get("ten_mon", ""),
                "loai_hoc_phan": loai_hp,
                "so_tin_chi": row.get("tong_so_chi", 3),
                "diem_thong_thuong": tx_list,
                "diem_giua_ky": row.get("giua_ky"),
                "diem_cuoi_ky": row.get("diem_cuoi_ky"),
                "diem_thuc_hanh_hien_tai": th_list,
                "diem_thuc_hanh_tich_hop": th1 if loai_hp == "tich_hop" else None,
                "diem_tong_ket": diem_tong_ket,
                "diem_chu": diem_chu,
                "diem_he_4": map_letter_to_gpa(diem_chu),
                "status_canh_bao": status_canh_bao,
                "source": "databricks",
            })
        return records

    # 2. Fallback về local mock DB
    for key, record in MOCK_GOLD_DB.items():
        if isinstance(key, tuple):
            sid, mm = key
        else:
            parts = str(key).split("_", 1)
            sid, mm = (parts[0], parts[1]) if len(parts) == 2 else (str(key), "")
        if mm != ma_mon:
            continue
        student_name = ""
        for u in USERS_DB.values():
            if u.get("student_id") == sid or u.get("username") == sid:
                student_name = u.get("full_name", "")
                break
        records.append({
            "_key": str(key),
            "student_id": sid,
            "ten_sv": student_name,
            "ma_mon": mm,
            "ten_mon": record.get("ten_mon", ""),
            "loai_hoc_phan": record.get("loai_hoc_phan", ""),
            "so_tin_chi": record.get("so_tin_chi", 0),
            "diem_thong_thuong": record.get("diem_thong_thuong", []),
            "diem_giua_ky": record.get("diem_giua_ky"),
            "diem_cuoi_ky": record.get("diem_cuoi_ky"),
            "diem_thuc_hanh_hien_tai": record.get("diem_thuc_hanh_hien_tai", []),
            "diem_thuc_hanh_tich_hop": record.get("diem_thuc_hanh_tich_hop"),
            "diem_tong_ket": record.get("diem_tong_ket"),
            "diem_chu": record.get("diem_chu", ""),
            "diem_he_4": record.get("diem_he_4") if record.get("diem_he_4") is not None else map_letter_to_gpa(record.get("diem_chu")),
            "status_canh_bao": record.get("status_canh_bao", "An toan"),
            "source": record.get("source", "local"),
        })
    return records


# ── Endpoints ─────────────────────────────────────────────────

@router.get("/courses", summary="Danh sach mon giang vien phu trach")
def get_my_courses(current_user: UserOut = _lecturer_dep):
    """Lấy danh sách môn học mà giảng viên đang được phân công"""
    courses = _get_lecturer_courses(current_user)
    return {"courses": courses, "total": len(courses)}


@router.get("/grades/{ma_mon}", summary="Danh sach diem SV trong mon")
def get_grades_by_course(
    ma_mon: str,
    current_user: UserOut = _lecturer_dep
):
    """Lấy danh sách điểm tất cả sinh viên trong môn học"""
    # Kiểm tra GV có quyền với môn này không
    my_courses = _get_lecturer_courses(current_user)
    my_course_ids = [c["ma_mon"] for c in my_courses]

    # Nếu GV không có môn nào thì cho phép xem (demo mode) — còn nếu có phân công thì chỉ xem môn của mình
    if my_course_ids and ma_mon not in my_course_ids:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"Bạn không có quyền xem điểm môn {ma_mon}"
        )

    grades = _get_grades_for_course(ma_mon)
    course_info = next((c for c in COURSES_DB if c.get("id") == ma_mon), {})
    return {
        "ma_mon": ma_mon,
        "ten_mon": course_info.get("name", ma_mon),
        "grades": grades,
        "total": len(grades)
    }


@router.get("/grades", summary="Danh sach diem tat ca mon GV phu trach")
def get_all_my_grades(current_user: UserOut = _lecturer_dep):
    """Lấy tất cả điểm SV của tất cả các môn GV đang phụ trách"""
    my_courses = _get_lecturer_courses(current_user)
    all_grades = {}
    for course in my_courses:
        ma_mon = course["ma_mon"]
        all_grades[ma_mon] = {
            "info": course,
            "grades": _get_grades_for_course(ma_mon)
        }
    return all_grades


@router.put("/grades/{student_id}/{ma_mon}", summary="Chinh sua diem sinh vien")
def update_student_grade(
    student_id: str,
    ma_mon: str,
    body: GradeEditBody,
    current_user: UserOut = _lecturer_dep
):
    """Chỉnh sửa điểm thành phần của một sinh viên trong môn học"""
    # Tìm bản ghi
    record_key = (student_id, ma_mon)
    if record_key not in MOCK_GOLD_DB:
        record_key = None
        # fallback: search
        for k in MOCK_GOLD_DB:
            if isinstance(k, tuple) and k[0] == student_id and k[1] == ma_mon:
                record_key = k
                break

    if record_key is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Không tìm thấy bản ghi điểm: SV {student_id} – Môn {ma_mon}"
        )

    record = MOCK_GOLD_DB[record_key]
    old_snapshot = dict(record)

    # Cập nhật các trường điểm được cung cấp
    if body.diem_thong_thuong is not None:
        record["diem_thong_thuong"] = body.diem_thong_thuong
    if body.diem_giua_ky is not None:
        record["diem_giua_ky"] = body.diem_giua_ky
    if body.diem_cuoi_ky is not None:
        record["diem_cuoi_ky"] = body.diem_cuoi_ky
    if body.diem_thuc_hanh_hien_tai is not None:
        record["diem_thuc_hanh_hien_tai"] = body.diem_thuc_hanh_hien_tai
    if body.diem_thuc_hanh_tich_hop is not None:
        record["diem_thuc_hanh_tich_hop"] = body.diem_thuc_hanh_tich_hop

    loai_hp = record.get("loai_hoc_phan", "ly_thuyet")

    # Map các trường của môn tích hợp
    if loai_hp == "tich_hop":
        if body.diem_thong_thuong is not None:
            record["diem_thuong_ky_lt_list"] = body.diem_thong_thuong
        if body.diem_giua_ky is not None:
            record["diem_giua_ky_lt"] = body.diem_giua_ky
        if body.diem_thuc_hanh_hien_tai is not None:
            record["diem_thuc_hanh_hien_tai"] = body.diem_thuc_hanh_hien_tai
            if len(body.diem_thuc_hanh_hien_tai) > 0:
                record["diem_thuc_hanh_tich_hop"] = body.diem_thuc_hanh_hien_tai[0]

    # get_grade_letter is defined globally

    ck = record.get("diem_cuoi_ky")
    
    if loai_hp == "ly_thuyet":
        tk_list = record.get("diem_thong_thuong") or []
        tk_avg = sum(tk_list) / len(tk_list) if tk_list else 0.0
        gk = record.get("diem_giua_ky")
        gk_val = gk if gk is not None else 0.0
        
        if ck is None:
            record["diem_tong_ket"] = None
            record["diem_chu"] = None
            record["diem_he_4"] = None
            record["ket_qua"] = None
            record["status_canh_bao"] = "An toan" if (0.2 * tk_avg + 0.3 * gk_val) >= 2.0 else "Nguy co"
        else:
            tong = round(0.2 * tk_avg + 0.3 * gk_val + 0.5 * ck, 1)
            tong = min(10.0, tong)
            record["diem_tong_ket"] = tong
            if ck < 3.0 or tong < 4.0:
                record["diem_chu"] = "F"
                record["diem_he_4"] = 0.0
                record["ket_qua"] = "Khong dat"
                record["status_canh_bao"] = "Nguy co"
            else:
                diem_chu, diem_he_4 = get_grade_letter(tong)
                record["diem_chu"] = diem_chu
                record["diem_he_4"] = diem_he_4
                record["ket_qua"] = "Dat"
                record["status_canh_bao"] = "An toan"
                
    elif loai_hp == "thuc_hanh":
        th_list = record.get("diem_thuc_hanh_hien_tai") or []
        avg_th = round(sum(th_list) / len(th_list), 1) if th_list else 0.0
        
        if ck is None:
            record["diem_tong_ket"] = None
            record["diem_chu"] = None
            record["diem_he_4"] = None
            record["ket_qua"] = None
            record["status_canh_bao"] = "An toan" if avg_th >= 4.0 else "Nguy co"
        else:
            record["diem_tong_ket"] = avg_th
            if avg_th < 3.0 or avg_th < 4.0:
                record["diem_chu"] = "F"
                record["diem_he_4"] = 0.0
                record["ket_qua"] = "Khong dat"
                record["status_canh_bao"] = "Nguy co"
            else:
                diem_chu, diem_he_4 = get_grade_letter(avg_th)
                record["diem_chu"] = diem_chu
                record["diem_he_4"] = diem_he_4
                record["ket_qua"] = "Dat"
                record["status_canh_bao"] = "An toan"
                
    else:  # tich_hop
        th_val = record.get("diem_thuc_hanh_tich_hop")
        th_val = th_val if th_val is not None else 0.0
        tk_lt_list = record.get("diem_thuong_ky_lt_list") or []
        tk_lt_avg = sum(tk_lt_list) / len(tk_lt_list) if tk_lt_list else 0.0
        gk_lt = record.get("diem_giua_ky_lt")
        gk_lt_val = gk_lt if gk_lt is not None else 0.0
        
        if ck is None:
            record["diem_tong_ket"] = None
            record["diem_chu"] = None
            record["diem_he_4"] = None
            record["ket_qua"] = None
            record["status_canh_bao"] = "An toan" if th_val >= 4.0 else "Nguy co"
        else:
            lt_score = 0.2 * tk_lt_avg + 0.3 * gk_lt_val + 0.5 * ck
            tong = round((lt_score * 2.0 + th_val * 1.0) / 3.0, 1)
            tong = min(10.0, tong)
            record["diem_tong_ket"] = tong
            if ck < 3.0 or th_val < 3.0 or tong < 4.0:
                record["diem_chu"] = "F"
                record["diem_he_4"] = 0.0
                record["ket_qua"] = "Khong dat"
                record["status_canh_bao"] = "Nguy co"
            else:
                diem_chu, diem_he_4 = get_grade_letter(tong)
                record["diem_chu"] = diem_chu
                record["diem_he_4"] = diem_he_4
                record["ket_qua"] = "Dat"
                record["status_canh_bao"] = "An toan"

    # Lưu lại MOCK_GOLD_DB
    MOCK_GOLD_DB[record_key] = record

    # Đồng bộ local silver db
    from app.db.databricks_db import sync_gold_to_silver, update_grades_in_cloud
    sync_gold_to_silver()

    # Tính toán qt_10 để cập nhật Databricks SQL Warehouse
    if loai_hp == "ly_thuyet":
        tk_list = record.get("diem_thong_thuong") or []
        tk_avg = sum(tk_list) / len(tk_list) if tk_list else 0.0
        gk = record.get("diem_giua_ky")
        gk_val = gk if gk is not None else 0.0
        qt_10 = round(0.2 * tk_avg + 0.3 * gk_val, 2)
    elif loai_hp == "thuc_hanh":
        th_list = record.get("diem_thuc_hanh_hien_tai") or []
        qt_10 = round(sum(th_list) / len(th_list), 2) if th_list else 0.0
    else:  # tich_hop
        th_val = record.get("diem_thuc_hanh_tich_hop")
        th_val = th_val if th_val is not None else 0.0
        tk_lt_list = record.get("diem_thuong_ky_lt_list") or []
        tk_lt_avg = sum(tk_lt_list) / len(tk_lt_list) if tk_lt_list else 0.0
        gk_lt = record.get("diem_giua_ky_lt")
        gk_lt_val = gk_lt if gk_lt is not None else 0.0
        qt_10 = round(((0.2 * tk_lt_avg + 0.3 * gk_lt_val) * 2.0 + th_val * 1.0) / 3.0, 2)

    update_grades_in_cloud(
        student_id=student_id,
        ma_mon=ma_mon,
        diem_thong_thuong=record.get("diem_thong_thuong") or record.get("diem_thuong_ky_lt_list"),
        diem_giua_ky=record.get("diem_giua_ky") or record.get("diem_giua_ky_lt"),
        diem_cuoi_ky=record.get("diem_cuoi_ky"),
        diem_thuc_hanh_hien_tai=record.get("diem_thuc_hanh_hien_tai"),
        qt_10=qt_10,
        diem_tong_ket=record.get("diem_tong_ket"),
        diem_chu=record.get("diem_chu"),
        diem_he_4=record.get("diem_he_4"),
        status_canh_bao_final=record.get("status_canh_bao")
    )

    record["updated_at"] = _now_str()
    record["updated_by"] = current_user.email

    # Log hoạt động
    log_event = {
        "id": f"act_{uuid4().hex[:8]}",
        "actor_email": current_user.email,
        "actor_name": current_user.full_name,
        "action": "grade_edit",
        "subject_id": student_id,
        "subject_name": f"SV {student_id} – Môn {ma_mon}",
        "details": f"Lý do: {body.reason or 'Không ghi chú'}",
        "timestamp": _now_str(),
    }
    ACTIVITY_LOGS.append(log_event)

    # Ghi vào score history
    SCORE_HISTORY_DB.append({
        "event_id": log_event["id"],
        "student_id": student_id,
        "ma_mon": ma_mon,
        "actor": current_user.email,
        "action": "edit",
        "old": old_snapshot,
        "new": dict(record),
        "timestamp": _now_str(),
    })

    try:
        from app.db.persistence import save_db_to_disk
        save_db_to_disk()
    except Exception as e:
        logger.error(f"Loi khi sao luu database: {e}")

    return {
        "message": f"Đã cập nhật điểm SV {student_id} môn {ma_mon}",
        "updated_record": record,
        "log": log_event
    }


@router.delete("/grades/{student_id}/{ma_mon}", summary="Xoa ban ghi diem sinh vien")
def delete_student_grade(
    student_id: str,
    ma_mon: str,
    reason: str = Query("Giảng viên xóa bản ghi"),
    current_user: UserOut = _lecturer_dep
):
    """Xóa bản ghi điểm của một sinh viên trong môn học"""
    record_key = (student_id, ma_mon)
    if record_key not in MOCK_GOLD_DB:
        record_key = None
        for k in MOCK_GOLD_DB:
            if isinstance(k, tuple) and k[0] == student_id and k[1] == ma_mon:
                record_key = k
                break

    if record_key is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Không tìm thấy bản ghi: SV {student_id} – Môn {ma_mon}"
        )

    deleted_record = MOCK_GOLD_DB.pop(record_key)

    log_event = {
        "id": f"act_{uuid4().hex[:8]}",
        "actor_email": current_user.email,
        "actor_name": current_user.full_name,
        "action": "grade_delete",
        "subject_id": student_id,
        "subject_name": f"SV {student_id} – Môn {ma_mon}",
        "details": f"Lý do: {reason}",
        "timestamp": _now_str(),
    }
    ACTIVITY_LOGS.append(log_event)
    SCORE_HISTORY_DB.append({
        "event_id": log_event["id"],
        "student_id": student_id,
        "ma_mon": ma_mon,
        "actor": current_user.email,
        "action": "delete",
        "old": deleted_record,
        "new": None,
        "timestamp": _now_str(),
    })

    try:
        from app.db.persistence import save_db_to_disk
        save_db_to_disk()
    except Exception as e:
        logger.error(f"Loi khi sao luu database: {e}")

    return {"message": f"Đã xóa bản ghi điểm SV {student_id} môn {ma_mon}", "log": log_event}


@router.post("/upload/{ma_mon}", summary="Upload bang diem CSV/XLSX cho mon hoc")
async def upload_grades_for_course(
    ma_mon: str,
    file: UploadFile = File(...),
    current_user: UserOut = _lecturer_dep
):
    """Upload bảng điểm từ file CSV hoặc XLSX cho một môn học cụ thể"""
    filename = file.filename or ""
    file_ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if file_ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Chỉ chấp nhận file .csv hoặc .xlsx"
        )

    contents = await file.read()
    rows: List[Dict[str, Any]] = []

    if file_ext == "csv":
        try:
            text = contents.decode("utf-8-sig")
        except Exception:
            text = contents.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text))
        rows = [row for row in reader]
    else:
        # XLSX support
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Server chưa cài openpyxl để đọc XLSX. Hãy dùng file CSV."
            )
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Không thể đọc file XLSX: {str(e)}"
            )

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="File trống hoặc không có dòng dữ liệu"
        )

    imported = 0
    errors = []
    course_info = COURSES_DB.get(ma_mon, {})

    for i, row in enumerate(rows):
        # Chuẩn hóa key
        row = {k.strip().lower().replace(" ", "_"): str(v).strip() for k, v in row.items()}
        student_id = row.get("student_id") or row.get("mssv") or row.get("ma_sv") or ""
        if not student_id:
            errors.append(f"Dòng {i+2}: Thiếu student_id")
            continue

        key = (student_id, ma_mon)
        def _parse_float(val):
            try:
                return float(val) if val else None
            except:
                return None

        def _parse_list(val):
            if not val:
                return []
            try:
                sep = ";" if ";" in val else ","
                return [float(x.strip()) for x in val.split(sep) if x.strip()]
            except:
                return []

        record = {
            "student_id": student_id,
            "ma_mon": ma_mon,
            "ten_mon": course_info.get("name", ma_mon),
            "loai_hoc_phan": row.get("loai_hoc_phan") or course_info.get("type", "ly_thuyet"),
            "so_tin_chi": int(course_info.get("credits", 0)),
            "diem_thong_thuong": _parse_list(row.get("diem_thong_thuong") or row.get("diem_thuong_ky") or ""),
            "diem_giua_ky": _parse_float(row.get("diem_giua_ky")),
            "diem_cuoi_ky": _parse_float(row.get("diem_cuoi_ky")),
            "diem_thuc_hanh_hien_tai": _parse_list(row.get("diem_thuc_hanh_hien_tai") or ""),
            "diem_thuc_hanh_tich_hop": _parse_float(row.get("diem_thuc_hanh_tich_hop")),
            "source": "lecturer_upload",
            "uploaded_by": current_user.email,
            "uploaded_at": _now_str(),
        }
        MOCK_GOLD_DB[key] = record
        imported += 1

    # Log
    ACTIVITY_LOGS.append({
        "id": f"act_{uuid4().hex[:8]}",
        "actor_email": current_user.email,
        "actor_name": current_user.full_name,
        "action": "grade_upload",
        "subject_id": ma_mon,
        "subject_name": f"Upload điểm môn {ma_mon}",
        "details": f"File: {filename} – {imported} bản ghi",
        "timestamp": _now_str(),
    })

    try:
        from app.db.persistence import save_db_to_disk
        save_db_to_disk()
    except Exception as e:
        logger.error(f"Loi khi sao luu database: {e}")

    return {
        "message": f"Đã nhập {imported} bản ghi điểm cho môn {ma_mon}",
        "ma_mon": ma_mon,
        "imported": imported,
        "errors": errors[:10],
        "total_rows": len(rows),
    }
